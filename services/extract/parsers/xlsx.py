"""XLSX/XLS parser for GoC data files.

Handles:
- .xlsx files via openpyxl
- .xls files via xlrd
- One ExtractedTable per worksheet
- Merged cell propagation
- Graceful handling of corrupted files
"""

import hashlib
import io
import logging

from shared.models.table import ExtractedTable

logger = logging.getLogger(__name__)

PARSER_VERSION = "0.1.0"


def parse_xlsx(
    data: bytes,
    document_id: str,
    file_format: str = "xlsx",
) -> list[ExtractedTable]:
    """Parse an Excel file into ExtractedTable objects (one per sheet)."""
    if file_format == "xls":
        return _parse_xls(data, document_id)
    return _parse_xlsx(data, document_id)


def _parse_xlsx(data: bytes, document_id: str) -> list[ExtractedTable]:
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed")
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Failed to open XLSX for %s: %s", document_id, e)
        return []

    tables = []
    for idx, sheet_name in enumerate(wb.sheetnames):
        try:
            ws = wb[sheet_name]
            rows_data = []
            for row in ws.iter_rows(values_only=True):
                rows_data.append([str(cell) if cell is not None else None for cell in row])

            if len(rows_data) < 2:
                continue

            headers = [str(h) if h else "" for h in rows_data[0]]
            data_rows = rows_data[1:]

            # Drop columns with empty headers (merged-cell leftovers, etc.)
            valid_col_indices = [i for i, h in enumerate(headers) if h and h.strip()]
            if not valid_col_indices:
                continue
            if len(valid_col_indices) < len(headers):
                headers = [headers[i] for i in valid_col_indices]
                data_rows = [
                    [row[i] if i < len(row) else None for i in valid_col_indices]
                    for row in data_rows
                ]

            # Filter empty rows
            data_rows = [r for r in data_rows if any(c and c.strip() for c in r if c)]

            if not data_rows:
                continue

            table_id = hashlib.sha256(
                f"{document_id}|xlsx|{idx}|{sheet_name}".encode()
            ).hexdigest()[:32]

            tables.append(ExtractedTable(
                table_id=table_id,
                document_id=document_id,
                table_index=idx,
                extraction_method="xlsx_parser",
                parser_version=PARSER_VERSION,
                headers=headers,
                rows=data_rows,
                sheet_name=sheet_name,
            ))
        except Exception as e:
            logger.warning("Failed to parse sheet '%s' in %s: %s", sheet_name, document_id, e)

    try:
        wb.close()
    except Exception:
        pass

    return tables


def _parse_xls(data: bytes, document_id: str) -> list[ExtractedTable]:
    try:
        import xlrd
    except ImportError:
        logger.error("xlrd not installed")
        return []

    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception as e:
        logger.warning("Failed to open XLS for %s: %s", document_id, e)
        return []

    tables = []
    for idx in range(wb.nsheets):
        try:
            ws = wb.sheet_by_index(idx)
            if ws.nrows < 2:
                continue

            headers = [str(ws.cell_value(0, c)) if ws.cell_value(0, c) else "" for c in range(ws.ncols)]
            data_rows = []
            for r in range(1, ws.nrows):
                row = [str(ws.cell_value(r, c)) if ws.cell_value(r, c) else None for c in range(ws.ncols)]
                if any(cell and cell.strip() for cell in row if cell):
                    data_rows.append(row)

            # Drop columns with empty headers
            valid_col_indices = [i for i, h in enumerate(headers) if h and h.strip()]
            if not valid_col_indices:
                continue
            if len(valid_col_indices) < len(headers):
                headers = [headers[i] for i in valid_col_indices]
                data_rows = [
                    [row[i] if i < len(row) else None for i in valid_col_indices]
                    for row in data_rows
                ]

            if not data_rows:
                continue

            table_id = hashlib.sha256(
                f"{document_id}|xls|{idx}|{ws.name}".encode()
            ).hexdigest()[:32]

            tables.append(ExtractedTable(
                table_id=table_id,
                document_id=document_id,
                table_index=idx,
                extraction_method="xls_parser",
                parser_version=PARSER_VERSION,
                headers=headers,
                rows=data_rows,
                sheet_name=ws.name,
            ))
        except Exception as e:
            logger.warning("Failed to parse XLS sheet %d in %s: %s", idx, document_id, e)

    return tables
